# -*- coding: utf-8 -*-
"""
Archivo: excelgen.py
Descripción: Generación y actualización del Excel con dashboard, registro de incidencias y contador de faltas.
"""

import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.chart import PieChart, Reference
from openpyxl.utils import get_column_letter

EXCEL_PATH = os.path.join("data", "bitacoras.xlsx")


def autosize_sheet(ws, min_width=8):
    """Ajusta ancho de columnas basado en el contenido (aproximado)."""
    dims = {}
    for row in ws.iter_rows(values_only=True):
        for idx, cell in enumerate(row, start=1):
            if cell is None:
                length = 0
            else:
                length = len(str(cell))
            dims[idx] = max(dims.get(idx, 0), length)
    for idx, width in dims.items():
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = max(min_width, width + 2)


def inicializar_excel():
    """
    Crea el archivo Excel con las hojas necesarias y los encabezados correctos si no existe.
    """
    if not os.path.exists("data"):
        os.makedirs("data")

    if not os.path.exists(EXCEL_PATH):
        wb = Workbook()

        # Dashboard (página principal)
        ws_dash = wb.active
        ws_dash.title = "Dashboard"
        ws_dash["A1"] = "Dashboard de Incidencias"
        ws_dash["A1"].font = Font(size=14, bold=True)
        ws_dash["A1"].alignment = Alignment(horizontal="center")
        ws_dash.merge_cells("A1:D1")

        # Hoja de Incidencias (simplificada)
        ws_inc = wb.create_sheet("Incidencias")
        ws_inc.append([
            "Fecha", "Hora", "Lugar", "Gravedad", "Participantes", "Link al Documento"
        ])

        # Hoja para el Registro de Faltas con columnas de severidad
        ws_faltas = wb.create_sheet("Registro de Faltas")
        ws_faltas.append(["Alumno", "Total de Faltas", "Leve", "Moderada", "Grave"])

        autosize_sheet(ws_inc)
        autosize_sheet(ws_faltas)
        wb.save(EXCEL_PATH)


def registrar_incidencia(datos):
    """
    Registra una nueva incidencia en la hoja 'Incidencias'.
    """
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Incidencias"]

    ws.append([
        datos["fecha"], datos["hora"], datos["lugar"], datos["gravedad"],
        ", ".join(datos["participantes"]), datos.get("link", "")
    ])

    autosize_sheet(ws)
    wb.save(EXCEL_PATH)


def actualizar_dashboard():
    """
    Actualiza la hoja Dashboard con el resumen de gravedad.
    """
    wb = load_workbook(EXCEL_PATH)
    ws_dash = wb["Dashboard"]
    ws_inc = wb["Incidencias"]

    # Limpiar contenido anterior del dashboard (desde fila 3 en adelante)
    for row in ws_dash.iter_rows(min_row=3):
        for cell in row:
            cell.value = None
    ws_dash._charts = []

    # --- Contar incidencias por gravedad ---
    total_gravedad = {"Leve": 0, "Moderada": 0, "Grave": 0}
    for row in ws_inc.iter_rows(min_row=2, max_col=4, values_only=True):
        if not row or not row[0]:
            continue
        gravedad = row[3]
        if gravedad in total_gravedad:
            total_gravedad[gravedad] += 1

    # Escribir la tabla de resumen de gravedad
    ws_dash["A3"] = "Gravedad"
    ws_dash["B3"] = "Cantidad"
    ws_dash["A3"].font = Font(bold=True)
    ws_dash["B3"].font = Font(bold=True)

    fila = 4
    for g, val in total_gravedad.items():
        ws_dash[f"A{fila}"] = g
        ws_dash[f"B{fila}"] = val
        fila += 1

    # Crear el gráfico de pastel
    pie = PieChart()
    pie.title = "Distribución de Incidencias por Gravedad"
    data = Reference(ws_dash, min_col=2, min_row=3, max_row=fila - 1)
    labels = Reference(ws_dash, min_col=1, min_row=4, max_row=fila - 1)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    ws_dash.add_chart(pie, "D3")

    autosize_sheet(ws_dash)
    wb.save(EXCEL_PATH)


def obtener_incidencias():
    """
    Devuelve una lista de incidencias registradas en la hoja 'Incidencias'.
    """
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Incidencias"]
    incidencias = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:  # Si hay fecha, consideramos que es una incidencia válida
            incidencias.append(f"Fecha: {row[0]}, Lugar: {row[2]}, Gravedad: {row[3]}, Participantes: {row[4]}")
    return incidencias


def eliminar_incidencia(indice):
    """
    Elimina una incidencia de la hoja 'Incidencias' dado su índice.
    """
    wb = load_workbook(EXCEL_PATH)
    ws = wb["Incidencias"]

    row_to_delete = indice + 2  # Ajustar índice porque la fila 1 es el encabezado
    ws.delete_rows(row_to_delete)

    wb.save(EXCEL_PATH)
